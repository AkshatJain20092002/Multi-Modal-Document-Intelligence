"""
Flattens the final list of ValidationResults (detected question + extracted
answer + validation verdict) into the "Question Bank" deliverable — CSV or
a real .xlsx file.
"""

from __future__ import annotations

import csv
from pathlib import Path

from app.retrieval.validator import ValidationResult

_FIELDNAMES = [
    "document_id", "page_number", "question_number", "question_type", "marks_raw",
    "question_text", "options", "has_diagram_reference", "detection_confidence",
    "answer_status", "answer_text", "answer_confidence", "source_pages",
    "context_pages_used", "verdict", "reasons",
]


def _row(result: ValidationResult) -> dict:
    answered = result.answered_question
    question = answered.question
    return {
        "document_id": question.document_id,
        "page_number": question.page_number,
        "question_number": question.question_number or "",
        "question_type": question.question_type or "",
        "marks_raw": question.marks_raw or "",
        "question_text": question.question_text,
        "options": " | ".join(question.options),
        "has_diagram_reference": question.has_diagram_reference,
        "detection_confidence": round(question.confidence, 2),
        "answer_status": answered.answer_status,
        "answer_text": answered.answer_text or "",
        "answer_confidence": round(answered.confidence, 2),
        "source_pages": ",".join(str(p) for p in answered.source_pages),
        "context_pages_used": ",".join(str(p) for p in answered.context_pages_used),
        "verdict": result.verdict,
        "reasons": "; ".join(result.reasons),
    }


def write_question_bank_csv(results: list[ValidationResult], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=_FIELDNAMES)
        writer.writeheader()
        for result in results:
            writer.writerow(_row(result))


def write_question_bank_xlsx(results: list[ValidationResult], output_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Question Bank"
    sheet.append(_FIELDNAMES)

    for result in results:
        row = _row(result)
        sheet.append([row[field] for field in _FIELDNAMES])

    widths = [18, 6, 14, 16, 10, 55, 45, 10, 10, 12, 55, 10, 10, 14, 8, 35]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"

    workbook.save(output_path)
